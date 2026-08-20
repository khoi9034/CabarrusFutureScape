from __future__ import annotations

import csv
import io
import json
from datetime import UTC, date, datetime
from decimal import Decimal
from types import SimpleNamespace

import pytest
from fastapi import Request
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.main import app
from app.models import (
    AccelaPlanReviewClean,
    FemaFloodZoneClean,
    ParcelEnriched,
    ParcelZoningOverlayV2,
    PermitIntelligenceSegment,
    RealPropertyPermitClean,
    RealPropertyPermitParcelRelationship,
    SchoolReference,
    SchoolZone,
    ZoningJurisdictionalClean,
)
from app.models.parcel import Base
from app.product.models import audit_events, organizations, product_metadata, users
from app.product.principal import (
    ROLE_PERMISSIONS,
    Permission,
    ProductPrincipal,
    Role,
)
from app.product.router import (
    database_session,
    get_master_data_source_db,
    get_product_principal,
)


SOURCE_TABLES = [
    ParcelEnriched.__table__,
    ParcelZoningOverlayV2.__table__,
    RealPropertyPermitClean.__table__,
    PermitIntelligenceSegment.__table__,
    RealPropertyPermitParcelRelationship.__table__,
    AccelaPlanReviewClean.__table__,
    ZoningJurisdictionalClean.__table__,
    FemaFloodZoneClean.__table__,
    SchoolZone.__table__,
    SchoolReference.__table__,
]
ORGANIZATION_ID = "00000000-0000-0000-0000-000000000101"
USER_ID = "00000000-0000-0000-0000-000000000110"


@pytest.fixture
def master_data_harness():
    product_engine = _sqlite_engine()
    product_metadata.create_all(
        product_engine,
        tables=[organizations, users, audit_events],
    )
    product_sessions = sessionmaker(
        bind=product_engine,
        expire_on_commit=False,
        class_=Session,
    )
    with product_sessions.begin() as session:
        session.execute(
            organizations.insert().values(
                id=ORGANIZATION_ID,
                name="Master Data Test",
                slug="master-data-test",
            ),
        )
        session.execute(
            users.insert().values(
                id=USER_ID,
                organization_id=ORGANIZATION_ID,
                external_subject="master-data-test-user",
                display_name="Master Data Test User",
                roles=[Role.PLANNER.value],
            ),
        )

    source_base_engine = _sqlite_engine()
    source_engine = source_base_engine.execution_options(
        schema_translate_map={"public": None},
    )
    Base.metadata.create_all(source_engine, tables=SOURCE_TABLES)
    source_sessions = sessionmaker(
        bind=source_engine,
        expire_on_commit=False,
        class_=Session,
    )
    _seed_sources(source_sessions)

    source_statements: list[str] = []

    @event.listens_for(source_base_engine, "before_cursor_execute")
    def track_source_sql(_connection, _cursor, statement, _parameters, _context, _many):
        source_statements.append(statement)

    current = {
        "principal": ProductPrincipal(
            subject="master-data-test-user",
            roles=frozenset({Role.PLANNER}),
            organization_id=ORGANIZATION_ID,
            user_id=USER_ID,
            authenticated=True,
        ),
    }

    def override_product_session():
        session = product_sessions()
        try:
            yield session
            session.commit()
        except Exception:
            session.rollback()
            raise
        finally:
            session.close()

    def override_source_session():
        session = source_sessions()
        try:
            yield session
            session.rollback()
        finally:
            session.close()

    def override_principal(request: Request) -> ProductPrincipal:
        request.state.product_principal = current["principal"]
        return current["principal"]

    app.dependency_overrides[database_session] = override_product_session
    app.dependency_overrides[get_master_data_source_db] = override_source_session
    app.dependency_overrides[get_product_principal] = override_principal
    try:
        with TestClient(app, raise_server_exceptions=False) as client:
            yield SimpleNamespace(
                client=client,
                current=current,
                product_sessions=product_sessions,
                source_statements=source_statements,
            )
    finally:
        app.dependency_overrides.clear()
        source_base_engine.dispose()
        product_engine.dispose()


def test_catalog_and_values_expose_only_governed_fields(master_data_harness) -> None:
    client = master_data_harness.client
    response = client.get("/api/v1/master-data/datasets")

    assert response.status_code == 200, response.text
    datasets = {item["id"]: item for item in response.json()["data"]}
    assert set(datasets) == {
        "parcels",
        "permits",
        "addresses",
        "zoning",
        "flood",
        "schools",
    }
    assert datasets["parcels"]["record_count"] == 4
    assert datasets["permits"]["record_count"] == 5
    assert {key: datasets[key]["record_count"] for key in datasets if key not in {"parcels", "permits"}} == {
        "addresses": 2,
        "zoning": 2,
        "flood": 2,
        "schools": 2,
    }
    restricted = {
        "acctname1",
        "mailaddr1",
        "geometry",
        "classification_reason",
        "rules_version",
        "source_last_modified_at",
        "is_high_value",
        "owner_name",
        "attributes",
        "source_url",
        "exclusion_reason",
    }
    for dataset in datasets.values():
        field_ids = {field["id"] for field in dataset["fields"]}
        assert restricted.isdisjoint(field_ids)
        assert set(dataset["default_fields"]) <= field_ids
        assert dataset["status"] == "ready"
        assert dataset["governance"]["access_mode"] == "read_only"
        assert dataset["governance"]["derived_outputs_only"] is True
        assert dataset["governance"]["sensitivity"] == "public_planner_safe"
        expected_formats = ["csv", "xlsx", "geojson"] if dataset["spatial"] else ["csv", "xlsx"]
        assert dataset["supported_export_formats"] == expected_formats

    relationship = datasets["permits"]["relationships"][0]
    assert relationship["id"] == "permits_to_parcels"
    assert {field["id"] for field in relationship["output_fields"]} == {
        "parcel_pin14",
        "parcel_acreage",
        "parcel_market_value",
        "parcel_zoning_code",
    }

    values = client.get(
        "/api/v1/master-data/datasets/permits/values/permit_status",
        params={"q": "Iss", "limit": 10},
    )
    assert values.status_code == 200, values.text
    assert values.json()["data"] == {"values": ["Issued"]}
    assert all(
        statement.lstrip().upper().startswith("SELECT")
        for statement in master_data_harness.source_statements
    )


@pytest.mark.parametrize(
    ("dataset_id", "fields", "filters", "expected_ids"),
    [
        (
            "parcels",
            ["official_parcel_id"],
            [{"field": "market_value", "operator": "gte", "value": "300"}],
            ["P-003", "P-004"],
        ),
        (
            "permits",
            ["permit_id"],
            [{"field": "permit_status", "operator": "eq", "value": "Issued"}],
            ["permit-001", "permit-002", "permit-003", "permit-004"],
        ),
        (
            "permits",
            ["permit_id"],
            [{"field": "permit_date", "operator": "gte", "value": "2025-04-01"}],
            ["permit-004", "permit-005"],
        ),
        (
            "parcels",
            ["official_parcel_id"],
            [{"field": "official_parcel_id", "operator": "contains", "value": "002"}],
            ["P-002"],
        ),
    ],
)
def test_preview_filters_and_rejects_untrusted_identifiers(
    master_data_harness,
    dataset_id: str,
    fields: list[str],
    filters: list[dict[str, str]],
    expected_ids: list[str],
) -> None:
    client = master_data_harness.client
    response = client.post(
        f"/api/v1/master-data/datasets/{dataset_id}/preview",
        json=_preview_payload(fields, filters),
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["total"] == len(expected_ids)
    assert [row[fields[0]] for row in data["rows"]] == expected_ids

    if dataset_id != "parcels" or filters[0]["field"] != "official_parcel_id":
        return

    for invalid_dataset, invalid_field in (
        ("parcels", "acctname1"),
        ("parcels", "not_a_field"),
    ):
        denied = client.post(
            f"/api/v1/master-data/datasets/{invalid_dataset}/preview",
            json=_preview_payload([invalid_field]),
        )
        assert denied.status_code == 422, denied.text

    unknown_dataset = client.get("/api/v1/master-data/datasets/not_a_dataset")
    assert unknown_dataset.status_code == 404, unknown_dataset.text

    master_data_harness.source_statements.clear()
    sql_like = client.post(
        "/api/v1/master-data/datasets/parcels/preview",
        json=_preview_payload(
            ["official_parcel_id"],
            [{
                "field": "official_parcel_id",
                "operator": "contains",
                "value": "%' OR 1=1 --",
            }],
        ),
    )
    assert sql_like.status_code == 200, sql_like.text
    sql_like_data = sql_like.json()["data"]
    assert {key: sql_like_data[key] for key in ("page", "page_size", "rows", "total")} == {
        "page": 1,
        "page_size": 50,
        "rows": [],
        "total": 0,
    }
    assert all("OR 1=1" not in statement for statement in master_data_harness.source_statements)


def test_preview_pagination_count_empty_and_stable_sort(master_data_harness) -> None:
    client = master_data_harness.client
    payload = _preview_payload(
        ["permit_id", "permit_amount"],
        page_size=2,
        sort_field="permit_amount",
    )
    first = client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json=payload,
    )
    second = client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json={**payload, "page": 2},
    )
    repeated = client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json={**payload, "page": 2},
    )

    assert first.status_code == second.status_code == repeated.status_code == 200
    assert first.json()["data"]["total"] == 5
    assert [row["permit_id"] for row in first.json()["data"]["rows"]] == [
        "permit-001",
        "permit-002",
    ]
    assert second.json()["data"]["rows"] == repeated.json()["data"]["rows"]
    assert [row["permit_id"] for row in second.json()["data"]["rows"]] == [
        "permit-003",
        "permit-004",
    ]

    empty = client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json=_preview_payload(
            ["permit_id"],
            [{"field": "permit_status", "operator": "eq", "value": "Missing"}],
        ),
    )
    assert empty.status_code == 200
    assert empty.json()["data"]["total"] == 0
    assert empty.json()["data"]["rows"] == []

    oversized = client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json=_preview_payload(["permit_id"], page_size=101),
    )
    assert oversized.status_code == 422


@pytest.mark.parametrize(
    ("dataset_id", "field", "operator", "value", "result_field", "expected"),
    [
        ("addresses", "site_address", "contains", "Main", "address_id", 1),
        ("zoning", "jurisdiction", "eq", "Concord", "zoning_id", "Z-001"),
        ("flood", "flood_severity", "eq", "High", "flood_zone_id", 1),
        ("schools", "school_level", "eq", "Elementary", "zone_id", "SZ-001"),
    ],
)
def test_added_datasets_use_governed_filters_and_spatial_preview(
    master_data_harness,
    dataset_id: str,
    field: str,
    operator: str,
    value: str,
    result_field: str,
    expected: object,
) -> None:
    response = master_data_harness.client.post(
        f"/api/v1/master-data/datasets/{dataset_id}/preview",
        json=_preview_payload(
            [result_field, field],
            [{"field": field, "operator": operator, "value": value}],
        ),
    )

    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["rows"][0][result_field] == expected
    assert data["spatial"] is True
    assert data["crs"] == "EPSG:4326"
    assert data["feature_collection"]["type"] == "FeatureCollection"
    assert len(data["feature_collection"]["features"]) == 1
    assert data["lineage"]["source_datasets"] == [dataset_id]
    assert data["lineage"]["query_timestamp"]


def test_permit_parcel_join_preserves_relationships_geometry_and_geojson(
    master_data_harness,
) -> None:
    catalog = master_data_harness.client.get("/api/v1/master-data/datasets").json()["data"]
    permits = next(dataset for dataset in catalog if dataset["id"] == "permits")
    all_joined_fields = [field["id"] for field in permits["fields"]]
    all_joined_fields.extend(
        field["id"] for field in permits["relationships"][0]["output_fields"]
    )
    select_all = master_data_harness.client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json={
            **_preview_payload(all_joined_fields),
            "join": {
                "relationship_id": "permits_to_parcels",
                "attach_geometry": False,
            },
        },
    )
    assert len(all_joined_fields) == 19
    assert select_all.status_code == 200, select_all.text
    assert select_all.json()["data"]["field_ids"] == all_joined_fields

    request = {
        "fields": [
            "permit_id",
            "official_parcel_id",
            "parcel_acreage",
            "parcel_zoning_code",
        ],
        "filters": [],
        "sort_field": "permit_id",
        "sort_direction": "asc",
        "join": {
            "relationship_id": "permits_to_parcels",
            "attach_geometry": True,
        },
    }
    preview = master_data_harness.client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json={**request, "page": 1, "page_size": 100},
    )

    assert preview.status_code == 200, preview.text
    data = preview.json()["data"]
    assert data["total"] == 6
    assert data["spatial"] is True
    assert data["geometry_type"] == "MultiPolygon"
    assert data["join_statistics"] == {
        "relationship_id": "permits_to_parcels",
        "source_records": 5,
        "matched_records": 4,
        "unmatched_records": 1,
        "match_percentage": 80.0,
        "output_records": 6,
    }
    assert [row["permit_id"] for row in data["rows"]].count("permit-004") == 2
    assert len(data["feature_collection"]["features"]) == 6
    assert sum(
        feature["geometry"] is None
        for feature in data["feature_collection"]["features"]
    ) == 1
    assert data["lineage"]["source_datasets"] == ["permits", "parcels"]
    assert data["lineage"]["geometry_source"] == "parcels"

    exported = master_data_harness.client.post(
        "/api/v1/master-data/datasets/permits/export",
        json={**request, "format": "geojson"},
        headers={"X-Request-ID": "master-data-join-geojson"},
    )
    assert exported.status_code == 200, exported.text
    assert "application/geo+json" in exported.headers["content-type"]
    assert "cfs_permits_" in exported.headers["content-disposition"]
    exported_data = exported.json()
    assert exported_data["type"] == "FeatureCollection"
    assert len(exported_data["features"]) == 6

    with master_data_harness.product_sessions() as session:
        audit = session.execute(
            select(audit_events).where(
                audit_events.c.request_id == "master-data-join-geojson",
            ),
        ).mappings().one()
    assert audit["details"]["lineage"]["export_format"] == "geojson"
    assert audit["details"]["lineage"]["matched_count"] == 4
    assert audit["details"]["lineage"]["unmatched_count"] == 1


def test_geojson_and_relationship_allowlist_fail_closed(master_data_harness) -> None:
    client = master_data_harness.client
    nonspatial = client.post(
        "/api/v1/master-data/datasets/permits/export",
        json={"fields": ["permit_id"], "filters": [], "format": "geojson"},
    )
    joined_field_without_join = client.post(
        "/api/v1/master-data/datasets/permits/preview",
        json=_preview_payload(["permit_id", "parcel_acreage"]),
    )
    relationship_on_wrong_dataset = client.post(
        "/api/v1/master-data/datasets/parcels/preview",
        json={
            **_preview_payload(["official_parcel_id"]),
            "join": {
                "relationship_id": "permits_to_parcels",
                "attach_geometry": True,
            },
        },
    )

    assert nonspatial.status_code == 422
    assert joined_field_without_join.status_code == 422
    assert relationship_on_wrong_dataset.status_code == 422


def test_csv_and_xlsx_exports_are_typed_formula_safe_and_audited(
    master_data_harness,
) -> None:
    client = master_data_harness.client
    fields = ["permit_number", "permit_amount", "permit_date"]
    request = {
        "fields": fields,
        "filters": [
            {"field": "permit_status", "operator": "eq", "value": "Issued"},
        ],
        "sort_field": "permit_id",
        "sort_direction": "asc",
    }
    csv_response = client.post(
        "/api/v1/master-data/datasets/permits/export",
        json={**request, "format": "csv"},
        headers={"X-Request-ID": "master-data-csv"},
    )
    xlsx_response = client.post(
        "/api/v1/master-data/datasets/permits/export",
        json={**request, "format": "xlsx"},
        headers={"X-Request-ID": "master-data-xlsx"},
    )

    assert csv_response.status_code == xlsx_response.status_code == 200
    assert "no-store" in csv_response.headers["cache-control"]
    assert csv_response.headers["x-content-type-options"] == "nosniff"
    assert "text/csv" in csv_response.headers["content-type"]
    assert "cfs_permits_" in csv_response.headers["content-disposition"]
    csv_rows = list(csv.reader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert csv_rows[0] == ["Permit number", "Permit amount", "Permit date"]
    assert len(csv_rows) == 5
    assert [row[0] for row in csv_rows[1:]] == [
        "'=2+2",
        "'+SUM(1,1)",
        "'-1+2",
        "'@HYPERLINK(\"x\")",
    ]
    assert [Decimal(row[1]) for row in csv_rows[1:]] == [
        Decimal("100"),
        Decimal("200"),
        Decimal("200"),
        Decimal("400"),
    ]
    assert [date.fromisoformat(row[2]) for row in csv_rows[1:]] == [
        date(2025, 1, 1),
        date(2025, 2, 1),
        date(2025, 3, 1),
        date(2025, 4, 1),
    ]

    assert "no-store" in xlsx_response.headers["cache-control"]
    assert xlsx_response.headers["x-content-type-options"] == "nosniff"
    assert "spreadsheetml.sheet" in xlsx_response.headers["content-type"]
    workbook = load_workbook(io.BytesIO(xlsx_response.content), data_only=False)
    try:
        assert workbook.sheetnames == ["Permits"]
        worksheet = workbook["Permits"]
        assert worksheet.sheet_state == "visible"
        assert all(not dimension.hidden for dimension in worksheet.column_dimensions.values())
        rows = list(worksheet.iter_rows())
        assert [cell.value for cell in rows[0]] == [
            "Permit number",
            "Permit amount",
            "Permit date",
        ]
        assert len(rows) == 5
        assert [row[0].value for row in rows[1:]] == [row[0] for row in csv_rows[1:]]
        assert all(row[0].data_type == "s" for row in rows[1:])
        assert all(isinstance(row[1].value, (int, float)) for row in rows[1:])
        assert all(isinstance(row[2].value, (date, datetime)) for row in rows[1:])
    finally:
        workbook.close()

    with master_data_harness.product_sessions() as session:
        audit = session.execute(
            select(audit_events).where(
                audit_events.c.request_id == "master-data-xlsx",
            ),
        ).mappings().one()
    assert audit["organization_id"] == ORGANIZATION_ID
    assert audit["actor_user_id"] == USER_ID
    assert audit["action"] == "master_data_export"
    assert audit["object_type"] == "master_data_dataset"
    assert audit["object_id"] == "permits"
    expected_audit = {
        "field_ids": fields,
        "field_count": 3,
        "filters": [{"field": "permit_status", "operator": "eq"}],
        "record_count": 4,
        "format": "xlsx",
        "runtime_mode": "local",
        "request_id": "master-data-xlsx",
    }
    assert {
        key: audit["details"][key] for key in expected_audit
    } == expected_audit
    assert audit["details"]["join_statistics"] is None
    assert audit["details"]["lineage"]["export_format"] == "xlsx"
    assert audit["details"]["lineage"]["query_timestamp"]
    serialized_audit = json.dumps(audit["details"], sort_keys=True)
    assert "Issued" not in serialized_audit
    assert "=2+2" not in serialized_audit


def test_permissions_fail_before_source_query_and_export_cap_is_enforced(
    master_data_harness,
    monkeypatch,
) -> None:
    client = master_data_harness.client
    master_data_harness.source_statements.clear()
    master_data_harness.current["principal"] = ProductPrincipal(
        subject="viewer",
        roles=frozenset({Role.VIEWER}),
        organization_id=ORGANIZATION_ID,
        user_id=USER_ID,
        authenticated=True,
    )
    view_denied = client.get("/api/v1/master-data/datasets")

    monkeypatch.setitem(
        ROLE_PERMISSIONS,
        Role.PLANNER,
        ROLE_PERMISSIONS[Role.PLANNER] - {Permission.MASTER_DATA_EXPORT},
    )
    master_data_harness.current["principal"] = ProductPrincipal(
        subject="planner-without-export",
        roles=frozenset({Role.PLANNER}),
        organization_id=ORGANIZATION_ID,
        user_id=USER_ID,
        authenticated=True,
    )
    export_denied = client.post(
        "/api/v1/master-data/datasets/permits/export",
        json={
            "fields": ["permit_id"],
            "filters": [],
            "format": "csv",
            "sort_direction": "asc",
        },
    )

    assert view_denied.status_code == export_denied.status_code == 403
    assert master_data_harness.source_statements == []

    monkeypatch.setitem(
        ROLE_PERMISSIONS,
        Role.PLANNER,
        ROLE_PERMISSIONS[Role.PLANNER] | {Permission.MASTER_DATA_EXPORT},
    )
    monkeypatch.setattr("app.product.master_data.EXPORT_ROW_CAP", 1)
    capped = client.post(
        "/api/v1/master-data/datasets/permits/export",
        json={
            "fields": ["permit_id"],
            "filters": [],
            "format": "csv",
            "sort_direction": "asc",
        },
        headers={"X-Request-ID": "master-data-cap"},
    )
    assert capped.status_code == 422, capped.text
    with master_data_harness.product_sessions() as session:
        assert session.scalar(
            select(audit_events.c.id).where(
                audit_events.c.request_id == "master-data-cap",
            ),
        ) is None


def _preview_payload(
    fields: list[str],
    filters: list[dict[str, str]] | None = None,
    *,
    page_size: int = 50,
    sort_field: str | None = None,
) -> dict[str, object]:
    return {
        "fields": fields,
        "filters": filters or [],
        "page": 1,
        "page_size": page_size,
        "sort_field": sort_field,
        "sort_direction": "asc",
    }


def _sqlite_engine():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def enable_foreign_keys(connection, _record) -> None:
        connection.execute("PRAGMA foreign_keys=ON")
        connection.create_function("ST_AsGeoJSON", 2, lambda geometry, _digits: geometry)

    return engine


def _seed_sources(source_sessions) -> None:
    polygon = json.dumps(
        {
            "type": "MultiPolygon",
            "coordinates": [[[[0, 0], [1, 0], [1, 1], [0, 0]]]],
        },
    )
    parcel_rows = []
    parcel_zoning_rows = []
    for index in range(1, 5):
        parcel_id = f"P-{index:03d}"
        parcel_rows.append(
            {
                "official_parcel_id": parcel_id,
                "pin14": f"PIN-{index:03d}",
                "subdiv_name": "North" if index < 3 else "South",
                "nbh_name": "Fixture neighborhood",
                "acctname1": "Restricted owner",
                "mailaddr1": "Restricted address",
                "marketvalue_numeric": Decimal(index * 100),
                "assessedvalue_numeric": Decimal(index * 80),
                "landvalue_numeric": Decimal(index * 50),
                "buildingvalue_numeric": Decimal(index * 30),
                "parcel_area_acres_calc": float(index),
                "value_per_acre": 100.0,
                "parcel_quality_status": "Internal",
                "enriched_at": datetime(2025, index, 1, tzinfo=UTC),
                "geometry": polygon,
            },
        )
        parcel_zoning_rows.append(
            {
                "official_parcel_id": parcel_id,
                "zoning_jurisdiction_name": "Concord",
                "dominant_zoning_code_raw": "R" if index < 3 else "C",
                "dominant_zoning_general_normalized": (
                    "Residential" if index < 3 else "Commercial"
                ),
            },
        )

    formula_numbers = [
        "=2+2",
        "+SUM(1,1)",
        "-1+2",
        '@HYPERLINK("x")',
        "SAFE-005",
    ]
    permit_rows = []
    segment_rows = []
    relationship_rows = []
    for index, permit_number in enumerate(formula_numbers, start=1):
        permit_id = f"permit-{index:03d}"
        permit_rows.append(
            {
                "permit_id": permit_id,
                "permit_number": permit_number,
                "parcel_number": f"SOURCE-{index:03d}",
                "permit_date": date(2025, index, 1),
                "permit_type_normalized": (
                    "Residential" if index % 2 else "Commercial"
                ),
                "work_type_normalized": "New construction",
                "permit_status_normalized": "Issued" if index < 5 else "Complete",
                "permit_amount": Decimal([100, 200, 200, 400, 500][index - 1]),
                "source_last_modified_at": datetime(2025, index, 2, tzinfo=UTC),
                "transformed_at": datetime(2025, index, 3, tzinfo=UTC),
            },
        )
        segment_rows.append(
            {
                "permit_id": permit_id,
                "permit_segment": "Housing" if index % 2 else "Employment",
                "permit_growth_signal": "Elevated" if index < 3 else "Baseline",
                "development_domain": "Development",
                "permit_value_class": "Major" if index > 3 else "Standard",
                "permit_status_stage": "Open" if index < 5 else "Closed",
                "classification_reason": "Restricted model detail",
                "rules_version": "restricted-v1",
                "is_high_value": index > 3,
            },
        )
        if index < 5:
            relationship_rows.append(
                {
                    "relationship_id": f"relationship-{index:03d}",
                    "permit_id": permit_id,
                    "official_parcel_id": f"P-{index:03d}",
                },
            )
    relationship_rows.append(
        {
            "relationship_id": "relationship-004-extra",
            "permit_id": "permit-004",
            "official_parcel_id": "P-001",
        },
    )

    address_rows = [
        {
            "accela_plan_review_id": 1,
            "official_parcel_id": "P-001",
            "pin14": "PIN-001",
            "address": "100 Main Street",
            "review_type": "Site plan",
            "review_status": "Open",
            "file_date": date(2025, 1, 1),
            "current_context_only": True,
            "cleaned_at": datetime(2025, 6, 1, tzinfo=UTC),
            "geometry": json.dumps({"type": "Point", "coordinates": [0.5, 0.5]}),
        },
        {
            "accela_plan_review_id": 2,
            "official_parcel_id": "P-002",
            "pin14": "PIN-002",
            "address": "200 Oak Avenue",
            "review_type": "Subdivision",
            "review_status": "Complete",
            "file_date": date(2025, 2, 1),
            "current_context_only": True,
            "cleaned_at": datetime(2025, 6, 2, tzinfo=UTC),
            "geometry": json.dumps({"type": "Point", "coordinates": [1.5, 1.5]}),
        },
    ]
    zoning_feature_rows = [
        {
            "zoning_jurisdictional_id": "Z-001",
            "jurisdiction_name": "Concord",
            "zoning_code_raw": "R-1",
            "zoning_general_normalized": "Residential",
            "zoning_type_raw": "Base",
            "base_district_raw": "R-1",
            "conditional_raw": None,
            "transformed_at": datetime(2025, 7, 1, tzinfo=UTC),
            "geometry": polygon,
        },
        {
            "zoning_jurisdictional_id": "Z-002",
            "jurisdiction_name": "Harrisburg",
            "zoning_code_raw": "C-1",
            "zoning_general_normalized": "Commercial",
            "zoning_type_raw": "Base",
            "base_district_raw": "C-1",
            "conditional_raw": None,
            "transformed_at": datetime(2025, 7, 2, tzinfo=UTC),
            "geometry": polygon,
        },
    ]
    flood_rows = [
        {
            "flood_zone_internal_id": 1,
            "fld_ar_id": "FLD-001",
            "flood_zone_code": "AE",
            "flood_constraint_type": "Special flood hazard area",
            "flood_severity_class": "High",
            "source_layer": "NFHL",
            "transformed_at": datetime(2025, 8, 1, tzinfo=UTC),
            "geometry": polygon,
        },
        {
            "flood_zone_internal_id": 2,
            "fld_ar_id": "FLD-002",
            "flood_zone_code": "X",
            "flood_constraint_type": "Minimal flood hazard",
            "flood_severity_class": "Minimal",
            "source_layer": "NFHL",
            "transformed_at": datetime(2025, 8, 2, tzinfo=UTC),
            "geometry": polygon,
        },
    ]
    school_reference_rows = [
        {
            "school_reference_id": "SR-001",
            "school_type": "Public",
            "address": "1 School Lane",
        },
        {
            "school_reference_id": "SR-002",
            "school_type": "Public",
            "address": "2 School Lane",
        },
    ]
    school_zone_rows = [
        {
            "zone_id": "SZ-001",
            "school_name_raw": "Example Elementary",
            "school_level": "Elementary",
            "school_system": "Cabarrus County Schools",
            "matched_school_reference_id": "SR-001",
            "match_confidence": "high",
            "include_in_cfs_v1": True,
            "source_layer": "Elementary zones",
            "transformed_at": datetime(2025, 9, 1, tzinfo=UTC),
            "geometry": polygon,
        },
        {
            "zone_id": "SZ-002",
            "school_name_raw": "Example High",
            "school_level": "High",
            "school_system": "Cabarrus County Schools",
            "matched_school_reference_id": "SR-002",
            "match_confidence": "high",
            "include_in_cfs_v1": True,
            "source_layer": "High zones",
            "transformed_at": datetime(2025, 9, 2, tzinfo=UTC),
            "geometry": polygon,
        },
    ]

    with source_sessions.begin() as session:
        session.execute(ParcelEnriched.__table__.insert(), parcel_rows)
        session.execute(ParcelZoningOverlayV2.__table__.insert(), parcel_zoning_rows)
        session.execute(RealPropertyPermitClean.__table__.insert(), permit_rows)
        session.execute(PermitIntelligenceSegment.__table__.insert(), segment_rows)
        session.execute(
            RealPropertyPermitParcelRelationship.__table__.insert(),
            relationship_rows,
        )
        session.execute(AccelaPlanReviewClean.__table__.insert(), address_rows)
        session.execute(ZoningJurisdictionalClean.__table__.insert(), zoning_feature_rows)
        session.execute(FemaFloodZoneClean.__table__.insert(), flood_rows)
        session.execute(SchoolReference.__table__.insert(), school_reference_rows)
        session.execute(SchoolZone.__table__.insert(), school_zone_rows)
