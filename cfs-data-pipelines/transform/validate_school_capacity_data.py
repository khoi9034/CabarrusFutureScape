"""Validate future school capacity and enrollment source files.

Phase 8D readiness only. This module validates templates and future files
without fabricating enrollment, capacity, utilization, or score values.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

VALID_SCHOOL_LEVELS = {"elementary", "middle", "high", "review_required"}
UTILIZATION_TOLERANCE = Decimal("1.0")


@dataclass(frozen=True)
class DatasetSpec:
    name: str
    required_columns: tuple[str, ...]
    year_column: str | None
    numeric_columns: tuple[str, ...]
    utilization_column: str | None
    enrollment_column: str | None
    capacity_column: str | None
    duplicate_key_columns: tuple[str, ...]


DATASET_SPECS: dict[str, DatasetSpec] = {
    "capacity": DatasetSpec(
        name="capacity",
        required_columns=(
            "school_name",
            "school_name_normalized",
            "school_level",
            "school_system",
            "school_year",
            "functional_capacity",
            "current_enrollment",
            "available_seats",
            "utilization_pct",
            "capacity_status",
            "source_name",
            "source_url",
            "notes",
        ),
        year_column="school_year",
        numeric_columns=(
            "functional_capacity",
            "current_enrollment",
            "available_seats",
            "utilization_pct",
        ),
        utilization_column="utilization_pct",
        enrollment_column="current_enrollment",
        capacity_column="functional_capacity",
        duplicate_key_columns=(
            "school_name_normalized",
            "school_level",
            "school_system",
            "school_year",
        ),
    ),
    "enrollment_history": DatasetSpec(
        name="enrollment_history",
        required_columns=(
            "school_name",
            "school_name_normalized",
            "school_level",
            "school_system",
            "school_year",
            "total_enrollment",
            "source_name",
            "source_url",
            "notes",
        ),
        year_column="school_year",
        numeric_columns=("total_enrollment",),
        utilization_column=None,
        enrollment_column="total_enrollment",
        capacity_column=None,
        duplicate_key_columns=(
            "school_name_normalized",
            "school_level",
            "school_system",
            "school_year",
        ),
    ),
    "grade_enrollment": DatasetSpec(
        name="grade_enrollment",
        required_columns=(
            "school_name",
            "school_name_normalized",
            "school_level",
            "school_system",
            "school_year",
            "grade_level",
            "grade_enrollment",
            "grade_capacity",
            "grade_utilization_pct",
            "source_name",
            "source_url",
            "notes",
        ),
        year_column="school_year",
        numeric_columns=(
            "grade_enrollment",
            "grade_capacity",
            "grade_utilization_pct",
        ),
        utilization_column="grade_utilization_pct",
        enrollment_column="grade_enrollment",
        capacity_column="grade_capacity",
        duplicate_key_columns=(
            "school_name_normalized",
            "school_level",
            "school_system",
            "school_year",
            "grade_level",
        ),
    ),
    "projection": DatasetSpec(
        name="projection",
        required_columns=(
            "school_name",
            "school_name_normalized",
            "school_level",
            "school_system",
            "projection_year",
            "projected_enrollment",
            "projected_capacity",
            "projected_utilization_pct",
            "projection_method",
            "source_name",
            "source_url",
            "notes",
        ),
        year_column="projection_year",
        numeric_columns=(
            "projected_enrollment",
            "projected_capacity",
            "projected_utilization_pct",
        ),
        utilization_column="projected_utilization_pct",
        enrollment_column="projected_enrollment",
        capacity_column="projected_capacity",
        duplicate_key_columns=(
            "school_name_normalized",
            "school_level",
            "school_system",
            "projection_year",
        ),
    ),
    "planned_capacity": DatasetSpec(
        name="planned_capacity",
        required_columns=(
            "school_name",
            "school_name_normalized",
            "school_level",
            "school_system",
            "project_name",
            "project_type",
            "planned_capacity_added",
            "planned_capacity_removed",
            "net_capacity_change",
            "expected_open_year",
            "status",
            "source_name",
            "source_url",
            "notes",
        ),
        year_column="expected_open_year",
        numeric_columns=(
            "planned_capacity_added",
            "planned_capacity_removed",
            "net_capacity_change",
        ),
        utilization_column=None,
        enrollment_column=None,
        capacity_column=None,
        duplicate_key_columns=(
            "school_name_normalized",
            "school_level",
            "school_system",
            "project_name",
        ),
    ),
}


def normalize_school_name(value: str | None) -> str:
    """Normalize school names to the Phase 8 school_reference convention."""

    if not value:
        return ""

    text = value.lower().strip()
    text = re.sub(r"\bschool\b", "", text)
    text = re.sub(r"\belem(?:entary)?\b|\bes\b", "elementary", text)
    text = re.sub(r"\bmiddle\b|\bms\b", "middle", text)
    text = re.sub(r"\bhigh\b|\bhs\b", "high", text)
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return Decimal(text.replace(",", ""))
    except InvalidOperation:
        return None


def parse_int(value: Any) -> int | None:
    decimal_value = parse_decimal(value)
    if decimal_value is None or decimal_value != decimal_value.to_integral_value():
        return None
    return int(decimal_value)


def calculate_utilization(enrollment: Any, capacity: Any) -> Decimal | None:
    enrollment_value = parse_decimal(enrollment)
    capacity_value = parse_decimal(capacity)
    if enrollment_value is None or capacity_value is None or capacity_value <= 0:
        return None
    return (enrollment_value / capacity_value * Decimal("100")).quantize(
        Decimal("0.0001")
    )


def classify_capacity_status(enrollment: Any, capacity: Any) -> str:
    utilization = calculate_utilization(enrollment, capacity)
    if utilization is None:
        return "not_available"
    if utilization < Decimal("85"):
        return "under_capacity"
    if utilization < Decimal("100"):
        return "near_capacity"
    if utilization < Decimal("115"):
        return "over_capacity"
    return "severely_over_capacity"


def _issue(
    *,
    dataset_name: str,
    source_file: str | None,
    row_number: int | None,
    row: dict[str, Any] | None,
    issue_type: str,
    severity: str,
    issue_description: str,
    recommended_fix: str,
) -> dict[str, Any]:
    row = row or {}
    return {
        "dataset_name": dataset_name,
        "source_file": source_file,
        "school_name": row.get("school_name"),
        "school_name_normalized": row.get("school_name_normalized"),
        "school_level": row.get("school_level"),
        "school_system": row.get("school_system"),
        "issue_type": issue_type,
        "severity": severity,
        "issue_description": issue_description,
        "recommended_fix": recommended_fix,
        "row_number": row_number,
    }


def validate_records(
    rows: list[dict[str, Any]],
    dataset_name: str,
    *,
    source_file: str | None = None,
    reference_lookup: set[tuple[str, str, str]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Validate source rows and return standardized rows plus QA issues."""

    if dataset_name not in DATASET_SPECS:
        raise ValueError(f"Unsupported dataset: {dataset_name}")

    spec = DATASET_SPECS[dataset_name]
    issues: list[dict[str, Any]] = []
    standardized: list[dict[str, Any]] = []

    if not rows:
        return [], [
            _issue(
                dataset_name=dataset_name,
                source_file=source_file,
                row_number=None,
                row=None,
                issue_type="empty_source_file",
                severity="error",
                issue_description="Source file contains no data rows.",
                recommended_fix="Provide a non-empty source file.",
            )
        ]

    actual_columns = set(rows[0].keys())
    missing_columns = [col for col in spec.required_columns if col not in actual_columns]
    for column in missing_columns:
        issues.append(
            _issue(
                dataset_name=dataset_name,
                source_file=source_file,
                row_number=None,
                row=None,
                issue_type="missing_required_column",
                severity="error",
                issue_description=f"Required column '{column}' is missing.",
                recommended_fix="Use the matching data_templates/schools CSV template.",
            )
        )

    seen_keys: dict[tuple[str, ...], int] = {}
    for index, raw_row in enumerate(rows, start=2):
        row = {key: ("" if value is None else str(value).strip()) for key, value in raw_row.items()}
        if not row.get("school_name_normalized"):
            row["school_name_normalized"] = normalize_school_name(row.get("school_name"))
        row["school_level"] = row.get("school_level", "").strip().lower()
        row["school_system"] = row.get("school_system", "").strip().upper()

        if row["school_level"] not in VALID_SCHOOL_LEVELS:
            issues.append(
                _issue(
                    dataset_name=dataset_name,
                    source_file=source_file,
                    row_number=index,
                    row=row,
                    issue_type="invalid_school_level",
                    severity="error",
                    issue_description="school_level must be elementary, middle, high, or review_required.",
                    recommended_fix="Correct school_level or mark uncertain rows as review_required.",
                )
            )

        if row["school_system"] and row["school_system"] != "CCS":
            issues.append(
                _issue(
                    dataset_name=dataset_name,
                    source_file=source_file,
                    row_number=index,
                    row=row,
                    issue_type="non_ccs_school_system",
                    severity="info",
                    issue_description="Non-CCS school system is preserved for future use but outside CFS V1 capacity scoring.",
                    recommended_fix="Confirm whether this row should remain outside CFS V1 scope.",
                )
            )

        if spec.year_column:
            year_value = parse_int(row.get(spec.year_column))
            if year_value is None or year_value < 1900 or year_value > 2100:
                issues.append(
                    _issue(
                        dataset_name=dataset_name,
                        source_file=source_file,
                        row_number=index,
                        row=row,
                        issue_type="invalid_year",
                        severity="error",
                        issue_description=f"{spec.year_column} must be an integer year.",
                        recommended_fix="Provide a four-digit school/projection year.",
                    )
                )

        for column in spec.numeric_columns:
            value = parse_decimal(row.get(column))
            if row.get(column, "") == "":
                issue_type = "missing_capacity_value" if "capacity" in column else "missing_enrollment_value"
                issues.append(
                    _issue(
                        dataset_name=dataset_name,
                        source_file=source_file,
                        row_number=index,
                        row=row,
                        issue_type=issue_type,
                        severity="review",
                        issue_description=f"{column} is missing and will remain unavailable.",
                        recommended_fix="Provide vetted value or leave explicitly blank for not_available.",
                    )
                )
            elif value is None:
                issues.append(
                    _issue(
                        dataset_name=dataset_name,
                        source_file=source_file,
                        row_number=index,
                        row=row,
                        issue_type="invalid_numeric_value",
                        severity="error",
                        issue_description=f"{column} must be numeric when supplied.",
                        recommended_fix="Remove formatting or provide a numeric value.",
                    )
                )
            elif value < 0:
                issues.append(
                    _issue(
                        dataset_name=dataset_name,
                        source_file=source_file,
                        row_number=index,
                        row=row,
                        issue_type="negative_numeric_value",
                        severity="error",
                        issue_description=f"{column} cannot be negative.",
                        recommended_fix="Correct the source value.",
                    )
                )

        if spec.enrollment_column and spec.capacity_column:
            enrollment = parse_decimal(row.get(spec.enrollment_column))
            capacity = parse_decimal(row.get(spec.capacity_column))
            if enrollment is not None and capacity is not None:
                if capacity < enrollment:
                    issues.append(
                        _issue(
                            dataset_name=dataset_name,
                            source_file=source_file,
                            row_number=index,
                            row=row,
                            issue_type="over_capacity",
                            severity="review",
                            issue_description="Enrollment exceeds capacity. This is allowed but should be reviewed.",
                            recommended_fix="Confirm source values with school capacity owner.",
                        )
                    )
                calculated = calculate_utilization(enrollment, capacity)
                provided = parse_decimal(row.get(spec.utilization_column)) if spec.utilization_column else None
                if calculated is not None and provided is not None and abs(calculated - provided) > UTILIZATION_TOLERANCE:
                    issues.append(
                        _issue(
                            dataset_name=dataset_name,
                            source_file=source_file,
                            row_number=index,
                            row=row,
                            issue_type="utilization_mismatch",
                            severity="review",
                            issue_description=f"Provided utilization {provided} differs from calculated utilization {calculated}.",
                            recommended_fix="Confirm whether utilization uses functional capacity or another denominator.",
                        )
                    )

        if reference_lookup is not None:
            key = (
                row.get("school_name_normalized", ""),
                row.get("school_level", ""),
                row.get("school_system", ""),
            )
            if all(key) and key not in reference_lookup:
                issues.append(
                    _issue(
                        dataset_name=dataset_name,
                        source_file=source_file,
                        row_number=index,
                        row=row,
                        issue_type="unmatched_school_reference",
                        severity="review",
                        issue_description="Row does not match public.school_reference by normalized name, level, and system.",
                        recommended_fix="Review school_name_normalized or add a governed reference alias if approved.",
                    )
                )

        key = tuple(row.get(column, "") for column in spec.duplicate_key_columns)
        if all(key):
            if key in seen_keys:
                issues.append(
                    _issue(
                        dataset_name=dataset_name,
                        source_file=source_file,
                        row_number=index,
                        row=row,
                        issue_type="duplicate_source_row",
                        severity="review",
                        issue_description=f"Duplicate key also appears on row {seen_keys[key]}.",
                        recommended_fix="Remove duplicate or clarify source revision.",
                    )
                )
            else:
                seen_keys[key] = index

        standardized.append(row)

    return standardized, issues


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on optional package
        raise RuntimeError("XLSX support requires openpyxl to be installed.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    return [
        {headers[index]: value for index, value in enumerate(row)}
        for row in rows
        if any(value is not None and str(value).strip() for value in row)
    ]


def read_tabular_file(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError("Unsupported file type. Use CSV or XLSX.")


def summarize_validation(rows: list[dict[str, Any]], issues: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "row_count": len(rows),
        "issue_count": len(issues),
        "error_count": sum(1 for issue in issues if issue["severity"] == "error"),
        "review_count": sum(1 for issue in issues if issue["severity"] == "review"),
        "info_count": sum(1 for issue in issues if issue["severity"] == "info"),
        "issues_by_type": {
            issue_type: sum(1 for issue in issues if issue["issue_type"] == issue_type)
            for issue_type in sorted({issue["issue_type"] for issue in issues})
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate future CFS school capacity/enrollment data files.",
    )
    parser.add_argument("--file", required=True, help="CSV or XLSX source file path.")
    parser.add_argument("--dataset", required=True, choices=sorted(DATASET_SPECS))
    parser.add_argument("--output-json", help="Optional path for validation JSON.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"Source file does not exist: {path}")

    rows = read_tabular_file(path)
    standardized, issues = validate_records(
        rows,
        args.dataset,
        source_file=str(path),
    )
    payload = {
        "generated_at": datetime.now().isoformat(),
        "dataset": args.dataset,
        "source_file": str(path),
        "summary": summarize_validation(standardized, issues),
        "issues": issues,
    }
    if args.output_json:
        Path(args.output_json).write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], indent=2))
    return 1 if payload["summary"]["error_count"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
